import json
import jsonpath
import requests
import openpyxl


def test_add_multiple_students():
    API_URL = 'http://thetestingworldapi.com/api/studentsDetails'
    file = open(
        "E:\PersonalHyd\Python\Programs\SeleniumInPython\TestCases\StudentDetails\TestCases\student_data.json", "r")
    wk = openpyxl.load_workbook(
        "E:\\PersonalHyd\\Python\\Programs\\SeleniumInPython\\TestCases\\DataDrivenTesting\\testData.xlsx")
    sh = wk['Sheet1']
    rows = sh.max_row
    json_request = json.loads(file.read())
    for i in range(2, rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value
        response = requests.post(API_URL, json_request)
        print(response.status_code)
        assert response.status_code == 201
        print(response.text)
